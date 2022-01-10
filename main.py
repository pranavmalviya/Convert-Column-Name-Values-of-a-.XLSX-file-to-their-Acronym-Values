from openpyxl import load_workbook
import pandas as pd
import re
#this code is written by Khushvardhan and Jayvardhan Bhardwaj
pd_xl_file = pd.ExcelFile("Sample.xlsx")
df = pd_xl_file.parse("Sheet1")
col_count=df.shape[1]
workbook = load_workbook(filename="Sample.xlsx")
sheet=workbook.active
from string import ascii_uppercase
import itertools

def iter_all_strings():
    for size in itertools.count(1):
        for s in itertools.product(ascii_uppercase, repeat=size):
            yield "".join(s)

s1=1
goodcol=""
while s1<col_count:
    for s in itertools.islice(iter_all_strings(),col_count):        
        cur_row = [s2 for s2 in re.split("([A-Z][^A-Z]*)", (sheet.cell(row=1,column=s1).value)) if s2]
        #this code is written by Khushvardhan and Jayvardhan Bhardwaj
        newcol=""
        if len(cur_row)==1 and "(*)" not in cur_row[0]:
            newcol=cur_row[0]
            goodcol=newcol
            s1+=1
            sheet[s+"1"]=newcol
            print(newcol)
        elif len(cur_row)==1 and "(*)" in cur_row[0]:
            
            newcol=cur_row[0].replace("(*)","")
            goodcol=newcol
            s1+=1
            sheet[s+"1"]=newcol
            print(newcol)
        elif cur_row[0][0]=="*":
            cur_row[0]=cur_row[0][1:]
            for j in range(0,len(cur_row)):
                if cur_row[j][0]==cur_row[j][0].upper():
                    newcol=newcol+cur_row[j][0]
                else:
                    continue
            #print(s)
            #print(type(sheet[s]))
            sheet[s+"1"]=newcol
            goodcol=newcol
            s1+=1
            print(newcol)
        elif cur_row[0][0]!="(":
            for j in range(0,len(cur_row)):
                if cur_row[j][0]==cur_row[j][0].upper():
                    newcol=newcol+cur_row[j][0]
                else:
                    continue
            sheet[s+"1"]=newcol
            goodcol=newcol
            s1+=1
        else:
            cur_row.pop(0)
            for j in range(0,len(cur_row)):
                if cur_row[j][0]==cur_row[j][0].upper():
                    newcol=newcol+cur_row[j][0]
                else:
                    continue
            newcol=goodcol+"_"+newcol
            sheet[s+"1"]=newcol
            s1+=1
            print(newcol)
workbook.save(filename="output.xlsx")
                
        
    
